# -*- coding: utf-8 -*-
import argparse
import io
import json
import os
import socketserver
import sys
import threading
import time
from datetime import datetime
from http.server import SimpleHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from zoneinfo import ZoneInfo

TZ_EKB = ZoneInfo("Asia/Yekaterinburg")


def now_ekb() -> str:
    return datetime.now(TZ_EKB).strftime('%Y-%m-%d %H:%M:%S')

import requests
from bs4 import BeautifulSoup
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = "user_status.db"
MONITOR_INTERVAL = 60


def get_db_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_database():
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            url TEXT NOT NULL,
            status TEXT NOT NULL,
            timestamp DATETIME NOT NULL,
            online_indicators TEXT,
            success BOOLEAN,
            error TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_url_timestamp ON user_status (url, timestamp)')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS tracked_urls (
            url TEXT PRIMARY KEY,
            interval_sec INTEGER NOT NULL DEFAULT 60,
            enabled BOOLEAN NOT NULL DEFAULT 1,
            last_status TEXT DEFAULT 'unknown',
            last_username TEXT,
            last_check DATETIME,
            added_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        conn.commit()
        conn.close()
        print(f"База данных инициализирована: {DB_PATH}")
    except sqlite3.DatabaseError:
        print(f"База данных повреждена, создаем новую...")
        os.remove(DB_PATH)
        init_database()
    except Exception as e:
        print(f"Ошибка при инициализации базы данных: {e}")


def save_status_to_db(result: dict) -> None:
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            url TEXT NOT NULL,
            status TEXT NOT NULL,
            timestamp DATETIME NOT NULL,
            online_indicators TEXT,
            success BOOLEAN,
            error TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        online_indicators_json = json.dumps(result.get('online_indicators', []), ensure_ascii=False)

        cursor.execute('''
        INSERT INTO user_status (url, status, timestamp, online_indicators, success, error)
        VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            result['url'],
            result['status'],
            result['timestamp'],
            online_indicators_json,
            result['success'],
            result.get('error')
        ))

        conn.commit()
        conn.close()
        print(f"Статус сохранен в базу данных: {result['url']} -> {result['status']}")
    except Exception as e:
        print(f"Ошибка при сохранении в базу данных: {e}")


def get_history_from_db(url: str, limit: int = 100, offset: int = 0) -> list:
    results = []
    try:
        conn = get_db_conn()
        cursor = conn.cursor()

        cursor.execute('''
        SELECT url, status, timestamp, online_indicators, success, error
        FROM user_status
        WHERE url = ?
        ORDER BY timestamp DESC
        LIMIT ? OFFSET ?
        ''', (url, limit, offset))

        rows = cursor.fetchall()
        for row in rows:
            entry = {
                'url': row['url'],
                'status': row['status'],
                'timestamp': row['timestamp'],
                'success': bool(row['success']),
            }
            if row['online_indicators']:
                try:
                    entry['online_indicators'] = json.loads(row['online_indicators'])
                except json.JSONDecodeError:
                    entry['online_indicators'] = []
            else:
                entry['online_indicators'] = []
            if row['error']:
                entry['error'] = row['error']
            results.append(entry)

        conn.close()
    except Exception as e:
        print(f"Ошибка при получении истории из базы данных: {e}")

    return results


def get_history_summary_from_db(url: str) -> dict:
    summary = {
        'total': 0, 'online': 0, 'offline': 0, 'unknown': 0,
        'first_check': None, 'last_check': None, 'url': url
    }
    try:
        conn = get_db_conn()
        cursor = conn.cursor()

        cursor.execute('''
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status = 'online' THEN 1 ELSE 0 END) as online,
            SUM(CASE WHEN status = 'offline' THEN 1 ELSE 0 END) as offline,
            SUM(CASE WHEN status = 'unknown' THEN 1 ELSE 0 END) as unknown,
            MIN(timestamp) as first_check,
            MAX(timestamp) as last_check
        FROM user_status
        WHERE url = ?
        ''', (url,))

        row = cursor.fetchone()
        if row:
            summary['total'] = row[0] or 0
            summary['online'] = row[1] or 0
            summary['offline'] = row[2] or 0
            summary['unknown'] = row[3] or 0
            summary['first_check'] = row[4]
            summary['last_check'] = row[5]

        conn.close()
    except Exception as e:
        print(f"Ошибка при получении сводки из базы данных: {e}")

    return summary


def get_all_urls_from_db() -> list:
    urls = []
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT url FROM user_status ORDER BY url')
        rows = cursor.fetchall()
        urls = [row[0] for row in rows]
        conn.close()
    except Exception as e:
        print(f"Ошибка при получении списка URL: {e}")
    return urls


def export_to_excel(url: str = None) -> bytes:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if url:
        rows = get_history_from_db(url, limit=100000)
        title = url
    else:
        urls = get_all_urls_from_db()
        rows = []
        for u in urls:
            rows.extend(get_history_from_db(u, limit=100000))
        rows.sort(key=lambda r: r['timestamp'], reverse=True)
        title = "Все пользователи"

    ws = wb.active
    ws.title = "Статус пользователей"

    headers = ["URL", "Статус", "Время проверки", "Индикаторы онлайн", "Успешно", "Ошибка"]
    col_widths = [50, 12, 22, 40, 10, 40]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    for row_idx, entry in enumerate(rows, 2):
        indicators = entry.get('online_indicators', [])
        indicators_str = '; '.join(
            f"{ind.get('selector', '')}: {ind.get('text', ind.get('content', ''))}"
            for ind in indicators
        ) if indicators else ''

        values = [
            entry['url'],
            entry['status'].upper(),
            entry['timestamp'],
            indicators_str,
            'Да' if entry.get('success') else 'Нет',
            entry.get('error', '')
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        status_cell = ws.cell(row=row_idx, column=2)
        if entry['status'] == 'online':
            status_cell.font = Font(color="008000", bold=True)
        elif entry['status'] == 'offline':
            status_cell.font = Font(color="FF0000", bold=True)
        else:
            status_cell.font = Font(color="FF8C00", bold=True)

    ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(rows) + 1}"
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_user_status(url: str, online_tags: list) -> dict:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers, timeout=20)
        response.raise_for_status()
        response.encoding = 'utf-8'

        soup = BeautifulSoup(response.content, 'html.parser')

        online_indicators = []
        for tag_config in online_tags:
            selector = tag_config.get('selector')
            text = tag_config.get('text', '')

            if selector:
                elements = soup.select(selector)
                for element in elements:
                    if text and text in element.get_text():
                        online_indicators.append({
                            'selector': selector,
                            'text': text,
                            'found': True
                        })
                    elif not text:
                        online_indicators.append({
                            'selector': selector,
                            'content': element.get_text().strip(),
                            'found': True
                        })

        is_online = len(online_indicators) > 0

        username = None
        title_tag = soup.select_one('h1.title')
        if title_tag:
            username = title_tag.get_text(strip=True)

        return {
            'status': 'online' if is_online else 'offline',
            'timestamp': now_ekb(),
            'url': url,
            'online_indicators': online_indicators,
            'success': True,
            'username': username
        }

    except requests.RequestException as e:
        return {
            'status': 'unknown',
            'error': f"Ошибка запроса: {str(e)}",
            'url': url,
            'timestamp': now_ekb(),
            'online_indicators': [],
            'success': False
        }
    except Exception as e:
        return {
            'status': 'unknown',
            'error': f"Ошибка парсинга: {str(e)}",
            'url': url,
            'timestamp': now_ekb(),
            'online_indicators': [],
            'success': False
        }


ONLINE_TAGS = [
    {'selector': 'div.top15', 'text': 'Жду звонка'},
    {'selector': 'div.free', 'text': 'Свободна'},
    {'selector': 'div.status-online', 'text': 'Онлайн'},
    {'selector': 'span.online-indicator', 'text': 'active'},
    {'selector': 'div.presence', 'text': 'Доступен'},
    {'selector': 'div.activity-status', 'text': 'Только что'}
]


# --- Server-side background monitoring ---

class BackgroundMonitor:
    def __init__(self):
        self._thread = None
        self._stop_event = threading.Event()

    def start(self):
        if self._thread and self._thread.is_alive():
            return
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        print("Фоновый мониторинг запущен")

    def stop(self):
        self._stop_event.set()
        print("Фоновый мониторинг остановлен")

    def _run(self):
        while not self._stop_event.is_set():
            try:
                conn = get_db_conn()
                cursor = conn.cursor()
                cursor.execute('SELECT url, interval_sec FROM tracked_urls WHERE enabled = 1')
                urls = cursor.fetchall()
                conn.close()

                for row in urls:
                    if self._stop_event.is_set():
                        break
                    url = row['url']
                    interval = row['interval_sec']
                    self._check_url(url)

                now = time.time()
                if urls:
                    next_check = min(
                        (now + row['interval_sec'] for row in urls),
                        default=now + 60
                    )
                    sleep_time = max(5, min(60, next_check - now))
                else:
                    sleep_time = 60

                self._stop_event.wait(sleep_time)

            except Exception as e:
                print(f"Ошибка в фоновом мониторинге: {e}")
                self._stop_event.wait(30)

    def _check_url(self, url: str):
        try:
            conn = get_db_conn()
            cursor = conn.cursor()
            cursor.execute('SELECT last_status FROM tracked_urls WHERE url = ?', (url,))
            row = cursor.fetchone()
            old_status = row['last_status'] if row else None
            conn.close()

            result = parse_user_status(url, ONLINE_TAGS)
            save_status_to_db(result)

            new_status = result['status']

            conn = get_db_conn()
            cursor = conn.cursor()
            cursor.execute('''
            UPDATE tracked_urls
            SET last_status = ?, last_username = ?, last_check = ?
            WHERE url = ?
            ''', (
                new_status,
                result.get('username'),
                now_ekb(),
                url
            ))
            conn.commit()
            conn.close()

            if old_status and old_status != new_status and new_status in ('online', 'offline'):
                send_telegram_notification(url, result.get('username'), old_status, new_status)
        except Exception as e:
            print(f"Ошибка при фоновой проверке {url}: {e}")


# --- Telegram notifications ---

TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '')
TELEGRAM_CHAT_ID = os.environ.get('TELEGRAM_CHAT_ID', '')


def send_telegram_notification(url: str, username: str | None, old_status: str, new_status: str) -> None:
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return

    emoji = '🟢' if new_status == 'online' else '🔴'
    status_text = 'В сети' if new_status == 'online' else 'Не в сети'
    display_name = username or url

    text = (
        f"{emoji} <b>Изменение статуса</b>\n"
        f"Пользователь: {display_name}\n"
        f"Статус: {status_text}\n"
        f"URL: {url}\n"
        f"Время: {now_ekb()}"
    )

    try:
        resp = requests.post(
            f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage',
            json={
                'chat_id': TELEGRAM_CHAT_ID,
                'text': text,
                'parse_mode': 'HTML',
                'disable_web_page_preview': True
            },
            timeout=10
        )
        if not resp.json().get('ok'):
            print(f"Ошибка Telegram: {resp.text}")
    except Exception as e:
        print(f"Ошибка отправки Telegram: {e}")


_monitor = BackgroundMonitor()


def add_tracked_url(url: str, interval: int = 60) -> bool:
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute('''
        INSERT OR REPLACE INTO tracked_urls (url, interval_sec, enabled, last_check)
        VALUES (?, ?, 1, ?)
        ''', (url, interval, now_ekb()))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Ошибка добавления URL в отслеживание: {e}")
        return False


def remove_tracked_url(url: str) -> bool:
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM tracked_urls WHERE url = ?', (url,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Ошибка удаления URL из отслеживания: {e}")
        return False


def get_tracked_urls() -> list:
    results = []
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM tracked_urls ORDER BY added_at')
        rows = cursor.fetchall()
        for row in rows:
            results.append({
                'url': row['url'],
                'interval_sec': row['interval_sec'],
                'enabled': bool(row['enabled']),
                'last_status': row['last_status'] or 'unknown',
                'last_username': row['last_username'],
                'last_check': row['last_check'],
                'added_at': row['added_at']
            })
        conn.close()
    except Exception as e:
        print(f"Ошибка получения списка отслеживаемых URL: {e}")
    return results


class SiteParserHandler(SimpleHTTPRequestHandler):
    def _send_json(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False, default=str).encode('utf-8'))

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def _read_body(self):
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length).decode('utf-8')
        return json.loads(body) if body else {}

    def do_POST(self):
        try:
            if self.path == '/parse':
                self._handle_parse()
            elif self.path == '/history':
                self._handle_history()
            elif self.path == '/track':
                self._handle_track()
            elif self.path == '/untrack':
                self._handle_untrack()
            else:
                self.send_error(404, "Not Found")
        except json.JSONDecodeError:
            self.send_error(400, "Неверный JSON")
        except Exception as exc:
            self._send_json({"error": f"Ошибка обработки: {exc}"}, 500)

    def _handle_parse(self):
        data = self._read_body()
        url = data.get('url', '').strip()
        if not url:
            self._send_json({"error": "URL не указан"}, 400)
            return

        result = parse_user_status(url, ONLINE_TAGS)
        save_status_to_db(result)
        self._send_json(result)

    def _handle_history(self):
        data = self._read_body()
        url = data.get('url', '').strip()
        if not url:
            self._send_json({"error": "URL не указан"}, 400)
            return

        limit = data.get('limit', 100)
        offset = data.get('offset', 0)
        include_summary = data.get('summary', True)

        history = get_history_from_db(url, limit, offset)
        response_data = {
            "success": True,
            "url": url,
            "history": history,
            "count": len(history),
            "limit": limit,
            "offset": offset
        }

        if include_summary:
            response_data["summary"] = get_history_summary_from_db(url)

        self._send_json(response_data)

    def _handle_track(self):
        data = self._read_body()
        url = data.get('url', '').strip()
        interval = data.get('interval', 60)

        if not url:
            self._send_json({"error": "URL не указан"}, 400)
            return

        success = add_tracked_url(url, interval)
        if success:
            _monitor.start()
        self._send_json({"success": success, "url": url, "tracking": True})

    def _handle_untrack(self):
        data = self._read_body()
        url = data.get('url', '').strip()

        if not url:
            self._send_json({"error": "URL не указан"}, 400)
            return

        success = remove_tracked_url(url)
        self._send_json({"success": success, "url": url, "tracking": False})

    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            try:
                with open('index.html', 'r', encoding='utf-8') as file:
                    content = file.read()
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(content.encode('utf-8'))
            except FileNotFoundError:
                self.send_error(404, "Файл index.html не найден")
        elif self.path == '/tracked':
            self._send_json({"success": True, "urls": get_tracked_urls()})
        elif self.path.startswith('/export-excel'):
            try:
                parsed = urlparse(self.path)
                params = parse_qs(parsed.query)
                url_filter = params.get('url', [None])[0]

                excel_data = export_to_excel(url=url_filter)

                filename = "user_status_all.xlsx"
                if url_filter:
                    safe_name = url_filter.replace('https://', '').replace('http://', '').replace('/', '_').replace(':', '_')
                    filename = f"user_status_{safe_name}.xlsx"

                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
                self.send_header('Content-Length', str(len(excel_data)))
                self.end_headers()
                self.wfile.write(excel_data)
            except Exception as exc:
                self.send_error(500, f"Ошибка экспорта: {exc}")
        else:
            self.send_error(404, "Not Found")

    def log_message(self, format, *args):
        return


def run_server(port: int = 8000) -> None:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    socketserver.TCPServer.allow_reuse_address = True

    _monitor.start()

    if TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID:
        print("Telegram-уведомления включены")
    else:
        print("Telegram-уведомления отключены (задайте TELEGRAM_BOT_TOKEN и TELEGRAM_CHAT_ID)")

    with socketserver.TCPServer(("", port), SiteParserHandler) as httpd:
        print(f"Сервер запущен: http://0.0.0.0:{port}/")
        print("Фоновый мониторинг запущен и будет работать после закрытия браузера.")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            pass
        finally:
            _monitor.stop()
            httpd.server_close()
            print("Сервер остановлен.")


def main() -> None:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')

    init_database()

    parser = argparse.ArgumentParser(description="Парсер статуса пользователя на сайте")
    parser.add_argument('--serve', action='store_true', help='Запустить веб-сервер с интерфейсом')
    parser.add_argument('--port', type=int, default=8000, help='Порт для веб-сервера (по умолчанию 8000)')
    parser.add_argument('url', nargs='?', help='URL профиля для проверки статуса')

    args = parser.parse_args()

    if args.serve:
        run_server(args.port)
        return

    if not args.url:
        parser.print_help()
        print("\nПримеры использования:")
        print("  python site_parser.py https://example.com/profile/123")
        print("  python site_parser.py --serve")
        print("  python site_parser.py --serve --port 3000")
        return

    result = parse_user_status(args.url, ONLINE_TAGS)
    save_status_to_db(result)

    if result['success']:
        status_ru = "в сети" if result['status'] == 'online' else "не в сети"
        print(f"Время проверки: {result['timestamp']}")
        print(f"Статус: {status_ru}")
        if result['online_indicators']:
            print("Найдены индикаторы онлайн:")
            for indicator in result['online_indicators']:
                text_info = indicator.get('text', indicator.get('content', ''))
                print(f"  - {indicator['selector']}: \"{text_info}\"")
    else:
        print(f"Ошибка: {result['error']}")


if __name__ == '__main__':
    main()
